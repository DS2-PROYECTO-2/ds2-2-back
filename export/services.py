import os
import tempfile
from io import BytesIO
from datetime import datetime, date
from django.conf import settings
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.core.files import File
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from users.models import User
from rooms.models import RoomEntry
from attendance.models import Attendance, Incapacity
from schedule.models import Schedule
from .models import ExportJob


class MonitorDataExporter:
    """
    Servicio para exportar datos de monitores en PDF y Excel
    """
    
    def __init__(self, export_job):
        self.export_job = export_job
        self.monitor_ids = export_job.monitor_ids
        self.start_date = export_job.start_date
        self.end_date = export_job.end_date
    
    def get_monitors_queryset(self):
        """Obtiene el queryset de monitores según los filtros"""
        queryset = User.objects.filter(role=User.MONITOR, is_verified=True)
        
        # Filtrar por IDs específicos si se proporcionan
        if self.monitor_ids:
            queryset = queryset.filter(id__in=self.monitor_ids)
        
        return queryset.order_by('first_name', 'last_name')
    
    def get_monitor_data(self, monitor):
        """Obtiene todos los datos relacionados con un monitor"""
        # Filtrar por fechas si se proporcionan
        date_filter = Q()
        if self.start_date:
            date_filter &= Q(entry_time__date__gte=self.start_date)
        if self.end_date:
            date_filter &= Q(entry_time__date__lte=self.end_date)
        
        # Datos de entradas a salas
        room_entries = monitor.room_entries.filter(date_filter).order_by('-entry_time')
        
        # Datos de turnos
        schedule_filter = Q()
        if self.start_date:
            schedule_filter &= Q(start_datetime__date__gte=self.start_date)
        if self.end_date:
            schedule_filter &= Q(start_datetime__date__lte=self.end_date)
        
        schedules = monitor.schedules.filter(schedule_filter).order_by('-start_datetime')
        
        # Datos de incapacidades
        incapacity_filter = Q()
        if self.start_date:
            incapacity_filter &= Q(start_date__gte=self.start_date)
        if self.end_date:
            incapacity_filter &= Q(end_date__lte=self.end_date)
        
        incapacities = monitor.incapacities.filter(incapacity_filter).order_by('-start_date')
        
        # Estadísticas
        total_hours = sum([entry.duration_hours for entry in room_entries if entry.duration_hours])
        total_entries = room_entries.count()
        total_schedules = schedules.count()
        total_incapacities = incapacities.count()
        
        return {
            'monitor': monitor,
            'room_entries': room_entries,
            'schedules': schedules,
            'incapacities': incapacities,
            'stats': {
                'total_hours': round(total_hours, 2),
                'total_entries': total_entries,
                'total_schedules': total_schedules,
                'total_incapacities': total_incapacities
            }
        }
    
    def export_to_pdf(self):
        """Exporta los datos a PDF usando BytesIO para evitar WinError 32"""
        try:
            # Crear buffer en memoria
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # Título del reporte
            title = Paragraph(f"Reporte de Monitores - {self.export_job.title}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Información del reporte
            report_info = [
                ['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M')],
                ['Período:', f"{self.start_date or 'Sin límite'} - {self.end_date or 'Sin límite'}"],
                ['Formato:', 'PDF'],
                ['Total de monitores:', str(self.get_monitors_queryset().count())]
            ]
            
            info_table = Table(report_info, colWidths=[2*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ]))
            
            story.append(info_table)
            story.append(Spacer(1, 30))
            
            # Datos de cada monitor
            monitors = self.get_monitors_queryset()
            for i, monitor in enumerate(monitors):
                if i > 0:
                    story.append(PageBreak())
                
                monitor_data = self.get_monitor_data(monitor)
                
                # Información del monitor
                monitor_title = Paragraph(f"Monitor: {monitor.get_full_name()}", heading_style)
                story.append(monitor_title)
                
                # Datos básicos del monitor
                basic_data = [
                    ['ID:', str(monitor.id)],
                    ['Nombre completo:', monitor.get_full_name()],
                    ['Identificación:', monitor.identification],
                    ['Email:', monitor.email],
                    ['Teléfono:', monitor.phone or 'No registrado'],
                    ['Verificado:', 'Sí' if monitor.is_verified else 'No'],
                    ['Fecha de registro:', monitor.created_at.strftime('%d/%m/%Y %H:%M')]
                ]
                
                basic_table = Table(basic_data, colWidths=[1.5*inch, 3*inch])
                basic_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                
                story.append(basic_table)
                story.append(Spacer(1, 20))
                
                # Estadísticas
                stats_title = Paragraph("Estadísticas", heading_style)
                story.append(stats_title)
                
                stats_data = [
                    ['Total de horas trabajadas:', f"{monitor_data['stats']['total_hours']} horas"],
                    ['Total de entradas a salas:', str(monitor_data['stats']['total_entries'])],
                    ['Total de turnos asignados:', str(monitor_data['stats']['total_schedules'])],
                    ['Total de incapacidades:', str(monitor_data['stats']['total_incapacities'])]
                ]
                
                stats_table = Table(stats_data, colWidths=[2.5*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                
                story.append(stats_table)
                story.append(Spacer(1, 20))
                
                # Entradas a salas recientes (últimas 10)
                if monitor_data['room_entries']:
                    entries_title = Paragraph("Entradas a Salas (Últimas 10)", heading_style)
                    story.append(entries_title)
                    
                    entries_data = [['Sala', 'Fecha Entrada', 'Fecha Salida', 'Duración (h)', 'Notas']]
                    for entry in monitor_data['room_entries'][:10]:
                        entries_data.append([
                            entry.room.name,
                            entry.entry_time.strftime('%d/%m/%Y %H:%M'),
                            entry.exit_time.strftime('%d/%m/%Y %H:%M') if entry.exit_time else 'En sala',
                            str(entry.duration_hours) if entry.duration_hours else 'N/A',
                            entry.notes[:50] + '...' if len(entry.notes) > 50 else entry.notes
                        ])
                    
                    entries_table = Table(entries_data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 0.8*inch, 1.5*inch])
                    entries_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ]))
                    
                    story.append(entries_table)
                    story.append(Spacer(1, 20))
            
            # Construir PDF en memoria
            doc.build(story)
            
            # Obtener contenido del buffer
            buffer.seek(0)
            pdf_content = buffer.getvalue()
            file_size = len(pdf_content)
            
            # Guardar archivo en el modelo usando BytesIO
            django_file = File(BytesIO(pdf_content))
            self.export_job.file.save(
                f"monitors_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                django_file,
                save=True
            )
            
            # Cerrar buffer
            buffer.close()
            
            # Marcar como completado
            self.export_job.mark_as_completed(file_size=file_size)
            
            return True
            
        except Exception as e:
            self.export_job.mark_as_failed(str(e))
            return False
    
    def export_to_excel(self):
        """Exporta los datos a Excel usando BytesIO para evitar WinError 32"""
        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Hoja de resumen
            ws_summary = wb.active
            ws_summary.title = "Resumen de Monitores"
            
            # Encabezados del resumen
            summary_headers = [
                'ID', 'Nombre Completo', 'Identificación', 'Email', 'Teléfono',
                'Verificado', 'Total Horas', 'Total Entradas', 'Total Turnos', 'Total Incapacidades',
                'Fecha Registro'
            ]
            
            for col, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Datos del resumen
            monitors = self.get_monitors_queryset()
            for row, monitor in enumerate(monitors, 2):
                monitor_data = self.get_monitor_data(monitor)
                
                ws_summary.cell(row=row, column=1, value=monitor.id)
                ws_summary.cell(row=row, column=2, value=monitor.get_full_name())
                ws_summary.cell(row=row, column=3, value=monitor.identification)
                ws_summary.cell(row=row, column=4, value=monitor.email)
                ws_summary.cell(row=row, column=5, value=monitor.phone or 'No registrado')
                ws_summary.cell(row=row, column=6, value='Sí' if monitor.is_verified else 'No')
                ws_summary.cell(row=row, column=7, value=monitor_data['stats']['total_hours'])
                ws_summary.cell(row=row, column=8, value=monitor_data['stats']['total_entries'])
                ws_summary.cell(row=row, column=9, value=monitor_data['stats']['total_schedules'])
                ws_summary.cell(row=row, column=10, value=monitor_data['stats']['total_incapacities'])
                ws_summary.cell(row=row, column=11, value=monitor.created_at.strftime('%d/%m/%Y %H:%M'))
            
            # Ajustar ancho de columnas
            for col in range(1, len(summary_headers) + 1):
                ws_summary.column_dimensions[get_column_letter(col)].width = 15
            
            # Hoja de entradas a salas
            ws_entries = wb.create_sheet("Entradas a Salas")
            
            entries_headers = [
                'ID Monitor', 'Monitor', 'Sala', 'Fecha Entrada', 'Fecha Salida',
                'Duración (h)', 'Activo', 'Notas'
            ]
            
            for col, header in enumerate(entries_headers, 1):
                cell = ws_entries.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Datos de entradas
            row = 2
            for monitor in monitors:
                monitor_data = self.get_monitor_data(monitor)
                for entry in monitor_data['room_entries']:
                    ws_entries.cell(row=row, column=1, value=monitor.id)
                    ws_entries.cell(row=row, column=2, value=monitor.get_full_name())
                    ws_entries.cell(row=row, column=3, value=entry.room.name)
                    ws_entries.cell(row=row, column=4, value=entry.entry_time.strftime('%d/%m/%Y %H:%M'))
                    ws_entries.cell(row=row, column=5, value=entry.exit_time.strftime('%d/%m/%Y %H:%M') if entry.exit_time else 'En sala')
                    ws_entries.cell(row=row, column=6, value=entry.duration_hours if entry.duration_hours else 0)
                    ws_entries.cell(row=row, column=7, value='Sí' if entry.active else 'No')
                    ws_entries.cell(row=row, column=8, value=entry.notes)
                    row += 1
            
            # Ajustar ancho de columnas
            for col in range(1, len(entries_headers) + 1):
                ws_entries.column_dimensions[get_column_letter(col)].width = 20
            
            # Hoja de turnos
            ws_schedules = wb.create_sheet("Turnos")
            
            schedules_headers = [
                'ID Monitor', 'Monitor', 'Sala', 'Fecha Inicio', 'Fecha Fin',
                'Duración (h)', 'Estado', 'Recurrente', 'Notas'
            ]
            
            for col, header in enumerate(schedules_headers, 1):
                cell = ws_schedules.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Datos de turnos
            row = 2
            for monitor in monitors:
                monitor_data = self.get_monitor_data(monitor)
                for schedule in monitor_data['schedules']:
                    ws_schedules.cell(row=row, column=1, value=monitor.id)
                    ws_schedules.cell(row=row, column=2, value=monitor.get_full_name())
                    ws_schedules.cell(row=row, column=3, value=schedule.room.name)
                    ws_schedules.cell(row=row, column=4, value=schedule.start_datetime.strftime('%d/%m/%Y %H:%M'))
                    ws_schedules.cell(row=row, column=5, value=schedule.end_datetime.strftime('%d/%m/%Y %H:%M'))
                    ws_schedules.cell(row=row, column=6, value=schedule.duration_hours)
                    ws_schedules.cell(row=row, column=7, value=schedule.get_status_display())
                    ws_schedules.cell(row=row, column=8, value='Sí' if schedule.recurring else 'No')
                    ws_schedules.cell(row=row, column=9, value=schedule.notes)
                    row += 1
            
            # Ajustar ancho de columnas
            for col in range(1, len(schedules_headers) + 1):
                ws_schedules.column_dimensions[get_column_letter(col)].width = 20
            
            # Guardar en buffer de memoria
            buffer = BytesIO()
            wb.save(buffer)
            
            # Obtener contenido del buffer
            buffer.seek(0)
            excel_content = buffer.getvalue()
            file_size = len(excel_content)
            
            # Guardar archivo en el modelo usando BytesIO
            django_file = File(BytesIO(excel_content))
            self.export_job.file.save(
                f"monitors_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                django_file,
                save=True
            )
            
            # Cerrar buffer
            buffer.close()
            
            # Marcar como completado
            self.export_job.mark_as_completed(file_size=file_size)
            
            return True
            
        except Exception as e:
            self.export_job.mark_as_failed(str(e))
            return False
